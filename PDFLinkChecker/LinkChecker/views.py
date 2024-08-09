from django.shortcuts import render, redirect
from django.db.utils import OperationalError
import pytz
from django.http import HttpResponse, HttpResponseRedirect
from django.http import HttpResponse
from django.http import JsonResponse
from .models import Links, Globals, Links_table, CheckLinkResult, User
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
import fitz
import requests
import glob
import datetime
import time
import os  
#import pandas as pd
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter
import smtplib,ssl
import threading
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from http.client import responses
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import views as auth_views
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.contrib import messages
from .forms import UserRegisterForm


PST = pytz.timezone('US/Pacific')
checkallLock = threading.Lock()

try:
    if (Globals.objects.all().count() == 0):
        Globals.objects.create(iteration=0, pdfDirectory="./pdf")

    globals = Globals.objects.first()
except: 
    pass  # happens when db doesn't exist yet, views.py should be
          # importable without this side effect

@login_required
def broken(request):
    return render(request, "LinkChecker/index.html", {
        #"links":Links_table.objects.filter(~Q(dismiss=True), ~Q(ignore = True)),
        "links":Links_table.objects.filter(broken=True),
        "allCount":Links_table.objects.all().count(),
        "brokenCount":Links_table.objects.filter(broken=True).count(),
        "dismissCount":Links_table.objects.filter(dismiss=True).count(),
        "ignoreCount":Links_table.objects.filter(ignore=True).count(),
        "globals": Globals.objects.first()
    })

@login_required
def all(request):
    return render(request, "LinkChecker/index.html", {
        "links":Links_table.objects.all(),
        "allCount":Links_table.objects.all().count(),
        "brokenCount":Links_table.objects.filter(broken=True).count(),
        "dismissCount":Links_table.objects.filter(dismiss=True).count(),
        "ignoreCount":Links_table.objects.filter(ignore=True).count(),
        "globals": Globals.objects.first()

    })

@login_required
def index(request):
    return broken(request)

@login_required
def dismiss(request):
    return render(request, "LinkChecker/index.html", {
        "links":Links_table.objects.filter(dismiss=True),
        "allCount":Links_table.objects.all().count(),
        "brokenCount":Links_table.objects.filter(broken=True).count(),
        "dismissCount":Links_table.objects.filter(dismiss=True).count(),
        "ignoreCount":Links_table.objects.filter(ignore=True).count(),
        "globals": Globals.objects.first()

    })

@login_required
def ignore(request):
    return render(request, "LinkChecker/index.html", {
        "links":Links_table.objects.filter(ignore=True),
        "allCount":Links_table.objects.all().count(),
        "brokenCount":Links_table.objects.filter(broken=True).count(),
        "dismissCount":Links_table.objects.filter(dismiss=True).count(),
        "ignoreCount":Links_table.objects.filter(ignore=True).count(),
        "globals": Globals.objects.first()

    })

@csrf_exempt
def dismissAction(request, id):
    obj = Links_table.objects.get(id = id)
    #obj.dismiss = True
    #obj.ignore = False
    #obj.save()
    Links_table.objects.filter(url=obj.url).update(broken=False, dismiss=True,ignore=False);
    return HttpResponse(status=200)

@csrf_exempt
def cancelDismissAction(request, id):
    try:
        obj = Links_table.objects.get(id = id)
        obj.broken = (obj.statusCode != 200)
        Links_table.objects.filter(url=obj.url).update(broken=obj.broken, dismiss=False,ignore=False);
        return JsonResponse({"broken": obj.broken})
    except:
        return HttpResponse(status=404)

@csrf_exempt
def ignoreAction(request, id):
    obj = Links_table.objects.get(id = id)
    #obj.ignore = True
    #obj.dismiss = False
    #obj.save()
    Links_table.objects.filter(url=obj.url).update(broken=False, dismiss=False,ignore=True);
    return HttpResponse(status=200)

@csrf_exempt
def cancelIgnoreAction(request, id):
    try:
        obj = Links_table.objects.get(id = id)
        obj.broken = (obj.statusCode != 200)
        Links_table.objects.filter(url=obj.url).update(broken=obj.broken, dismiss=False,ignore=False);
        return JsonResponse({"broken": obj.broken})
    except:
        return HttpResponse(status=404)


def get_all_pdfs():
    # TODO: handle multiple directories
    return list(glob.iglob(globals.pdfDirectory+"/*.pdf", recursive=True))

def get_first_link_instance(check_link, pdf):
    timesSeen = 0
    urlAnchor = ""
    doc = fitz.open(pdf)
    for page_num in range(doc.page_count):
        try: 
            page = doc.load_page(page_num)
        except:
            continue
    links = page.get_links()
    for link in links:
        if (link["kind"] == fitz.LINK_URI and (not link["uri"].startswith('mailto'))):
            url = link["uri"]
            if (url == check_link):
                timesSeen += 1
                if timesSeen == 1:
                    boundingBox = 10 
                    urlAnchor = page.get_textbox(link["from"] + (-boundingBox, -1, boundingBox, 1))
            if timesSeen == 2:
                break
            else:
                continue
        continue
    return timesSeen > 0, urlAnchor, timesSeen > 1

def get_all_links(pdf):
    # if multiple instances of a link exist, pick only the first and set moreInPdf to True
    uniqueLinks = {}
    doc = fitz.open(pdf)
    for page_num in range(doc.page_count):
        try: 
            page = doc.load_page(page_num)
        except:
            continue
    links = page.get_links()
    for link in links:
        if (link["kind"] == fitz.LINK_URI and (not link["uri"].startswith('mailto'))):
            url = link["uri"]
            # print(f'        Found url:{url}')
            if (url in uniqueLinks):
                uniqueLinks[url]['moreInPdf'] = True
                continue  # pick only the first
            else:
                uniqueLinks[url] = {"moreInPdf":False}
                boundingBox = 10 
                uniqueLinks[url]["linkText"] = page.get_textbox(link["from"] + (-boundingBox, -1, boundingBox, 1))
                # print(f'        unique_links: {uniqueLinks}')
                # TODO: if link anchor is an image, set linkText as "<image>"

    return uniqueLinks

def make_request(url):
    # TODO use proper request header and try different header options
    # TODO collect logs
    try: 
        response = requests.head(url, allow_redirects=True)   
    except:
        return 500, 'Internal Server Error', url
    
    final_url = response.url
    if (response.status_code != 200):
        if (response.status_code in [301, 302]):
            redirected_response = requests.head(final_url)
            if (redirected_response.status_code != 200):
                return redirected_response.status_code, redirected_response.reason, final_url

    return response.status_code, response.reason, response.url

def checkall_links():
    curIteration = globals.iteration + 1
    newBrokenLinksFound = False
    print(f'Processing iteration #{curIteration}')
    pdfs = get_all_pdfs()

    for pdf in pdfs:
        if (os.path.isdir(pdf)):
            continue
        # print(f'  Processing file {pdf}')
        print('+', end="", flush=True)
        links = get_all_links(pdf)

        for link in links:
            # print(f'    Processing url:{link}')
            print('.', end="", flush=True)
            linkObjs = Links_table.objects.filter(url=link)

            if (len(linkObjs) == 0):
                #there are no objects with this link
                statusCode, reason, finalUrl = make_request(link)
                # print(f'      statusCode:{statusCode}, reason:{reason}, finalUrl:{finalUrl}')
                newObj = Links_table.objects.create(url=link, statusCode=statusCode, reason=reason, pdfSource=pdf,finalUrl=finalUrl, 
                    lastChecked=datetime.datetime.now(PST),brokenSince=datetime.datetime.now(PST),urlText=links[link]["linkText"],moreInPdf=links[link]["moreInPdf"],lastIteration=curIteration)
                if (statusCode != 200):
                    newObj.broken = True
                    newBrokenLinksFound = True
                    newObj.brokenSince = newObj.lastChecked
                else:
                    newObj.broken = False
                newObj.save()
                continue  # to next link

            try:
                thisPdfObj = Links_table.objects.get(url=link, pdfSource=pdf)
            except:
                thisPdfObj = None
            
            if (thisPdfObj == None):
                linkObjsProcessed = linkObjs.filter(lastIteration__gte=curIteration)
                if (len(linkObjsProcessed) > 0):
                    # create new obj for this pdf with the same state at already processed
                    #already processed this link
                    linkObjsProcessed[0].pk = None
                    linkObjsProcessed[0]._state.adding = True
                    linkObjsProcessed[0].pdfSource = pdf
                    linkObjsProcessed[0].urlText=links[link]["linkText"]
                    linkObjsProcessed[0].moreInPdf=links[link]["moreInPdf"]
                    linkObjsProcessed[0].save()
                    continue # to next link
                """ else :
                    linkObjs[0].pk = None
                    linkObjs[0]._state.adding = True
                    linkObjs[0].pdfSource = pdf
                    linkObjs[0].urlText=links[link]["linkText"]
                    linkObjs[0].moreInPdf=links[link]["moreInPdf"]
                    linkObjs[0].save()
                continue # to next link """
            else:
                thisPdfObj.urlText = links[link]["linkText"]
                thisPdfObj.moreInPdf = links[link]["moreInPdf"]
                thisPdfObj.save();
        
            if (linkObjs.filter(ignore=True).count() > 0):
                if (thisPdfObj == None):
                    linkObjs[0].pk = None
                    linkObjs[0]._state.adding = True
                    linkObjs[0].pdfSource = pdf
                    linkObjs[0].urlText=links[link]["linkText"]
                    linkObjs[0].moreInPdf=links[link]["moreInPdf"]
                    linkObjs[0].save()
                Links_table.objects.filter(url=link).update(ignore=True, lastIteration=curIteration)
                continue  # to next link

            statusCode, reason, finalUrl = make_request(link)
            # print(f'      statusCode:{statusCode}, reason:{reason}, finalUrl:{finalUrl}')
            if (linkObjs.filter(statusCode=statusCode).exists()):
                if (thisPdfObj == None):
                    linkObjs[0].pk = None
                    linkObjs[0]._state.adding = True
                    linkObjs[0].pdfSource = pdf
                    linkObjs[0].urlText=links[link]["linkText"]
                    linkObjs[0].moreInPdf=links[link]["moreInPdf"]
                    linkObjs[0].statusCode = statusCode
                    linkObjs[0].reason = reason
                    linkObjs[0].finalUrl = finalUrl
                    linkObjs[0].save()
                Links_table.objects.filter(url=link).update(lastIteration=curIteration, lastChecked=datetime.datetime.now(PST))
            else:
                for obj in linkObjs:
                    if (statusCode != 200):
                        obj.broken = True
                        if (obj.statusCode == 200):
                            obj.broken_since = obj.last_checked
                    else: 
                        obj.broken = False
                    obj.dismiss = False #exit out of dismissed state
                    obj.save()

                if (thisPdfObj == None):
                    obj.pk = None
                    obj._state.adding = True
                    obj.pdfSource = pdf
                    obj.urlText=links[link]["linkText"]
                    obj.moreInPdf=links[link]["moreInPdf"]
                    obj.save()
                Links_table.objects.filter(url=link).update(statusCode=statusCode,reason=reason,finalUrl=finalUrl,lastIteration=curIteration, 
                    lastChecked=datetime.datetime.now(PST))
        # processed all links
    # processed all pdfs
    print(' Done!')
    #delete all stale links
    Links_table.objects.filter(lastIteration__lt=curIteration).delete()
    globals.iteration = curIteration
    globals.save()

    if (globals.emailNotifyOnNewLink and newBrokenLinksFound):
        print(f'sending email')
        if (globals.attachListToEmail):
            print(f'preparing attachment')
            attachmentFile = "PDFBrokenLinks.xlsx"
            # Create a new Excel workbook
            workbook = openpyxl.Workbook()
            # Select the default sheet (usually named 'Sheet')
            sheet = workbook.active
            # Add data to the Excel sheet
            data = [["URL", "Status Code", "Status Reason", "PDF Source"]] #top row
            for link in Links_table.objects.filter(broken=True):
                row = [link.url, link.statusCode, link.reason, link.pdfSource]
                data.append(row)
                #print(data)
            for row in data:
                sheet.append(row)
            # Save the workbook to a file
            workbook.save(attachmentFile)
            # Print a success message
            print("Excel file created successfully!")

        # send e-mail
        msg = MIMEMultipart()
        sender='murali.singamsetty@gmail,com'
        recipients= globals.emailAddress
        server=smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login("murali.singamsetty@gmail.com", "ysxyoczkwjzmswiu")
        msg['Subject']='PDF Broken Link Report'
        msg['From']=sender
        msg['To']=recipients
        mail_body = """\
            This is an automated report, sent from PDF Broken Link Checker.
            
            To disable email notifications, see click Settings on PDF Broken Link Checker

            """
        msg.attach(MIMEText(mail_body, "plain"))
        if (globals.attachListToEmail):
            print(f'adding attachment to email')
            attachment = open(attachmentFile, 'rb')
            xlsx = MIMEBase('application','vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            xlsx.set_payload(attachment.read())
            encoders.encode_base64(xlsx)
            xlsx.add_header('Content-Disposition', 'attachment', filename="PDFBrokenLinks.xlsx")
            msg.attach(xlsx)
        
        server.sendmail(sender, recipients, msg.as_string())
        print(f'sent email')
        server.quit()
        if (globals.attachListToEmail):
            attachment.close()

def download_excel(request):
    # Create an in-memory workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF Broken Links-" + datetime.datetime.now(PST).strftime("%m_%d_%Y")

    headers = ["URL", "Status Code", "Status Reason", "PDF Source"] 
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

    for row_num, obj in enumerate(Links_table.objects.filter(broken=True), 2):
        row = [obj.url, obj.statusCode, obj.reason, obj.pdfSource]  
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Adjust column widths
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Save the workbook to an in-memory file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="PDFBrokenLinks.xlsx"'
    wb.save(response)

    return response

def bgnd_task():

    while True:
        print('background task: acquiring lock')
        checkallLock.acquire(blocking=1)
        print(f'background task: got lock, processing at {datetime.datetime.now(PST)}')
        checkall_links()
        print('background task: releasing lock')
        checkallLock.release()
        next_at = datetime.datetime.now(PST)
        while (True):
            next_at = datetime.datetime(year=next_at.year, month=next_at.month, day=next_at.day, 
                        hour=globals.checkAllStartAtHour, minute=globals.checkAllStartAtMin).replace(tzinfo=PST) + datetime.timedelta(hours=globals.checkAllIntervalHours, 
                        minutes=globals.checkAllIntervalMins)

            if (next_at > datetime.datetime.now(PST)):
                break
            next_at += datetime.timedelta(hours=globals.checkAllIntervalHours, minutes=globals.checkAllIntervalMins)

        print(f'background task: scheduled to run again at {next_at}')
        total_seconds = (next_at - datetime.datetime.now(PST)).total_seconds()
        time.sleep(total_seconds)


def checkall(request):
    print('checkall request: acquiring lock')
    checkallLock.acquire(blocking=1)
    print(f'checkall request: got lock, processing at {datetime.datetime.now(PST)}')
    checkall_links()
    print('checkall request: releasing lock')
    checkallLock.release()
    return HttpResponseRedirect("/")

def settings(request, id):
    object = Globals.objects.first()
    if id == "directory":
        object.pdfDirectory = request.POST["directory"]
        object.save()
        print(object.pdfDirectory)
        return HttpResponseRedirect("/")
    if id =="time":
        object.checkAllStartAtHour = request.POST["hourStart"]
        object.checkAllStartAtMin = request.POST["minStart"]
        object.checkAllIntervalHours = request.POST["hourInterval"]
        object.checkAllIntervalMins = request.POST["minInterval"]
        object.save()
        return HttpResponseRedirect("/")
    if id =="emails":
        if "checkAttach" in request.POST:
            object.attachListToEmail = True
        else:
            object.attachListToEmail = False
        if "checkNotification" in request.POST:
            object.emailNotifyOnNewLink = True
        else:
            object.emailNotifyOnNewLink = False
        object.emailAddress = request.POST["emailadd"]
        object.save()
        return HttpResponseRedirect("/")


@csrf_exempt
def recheckAction(request, id):
    try:
        Obj = Links_table.objects.get(id = id)
    except:
        result =  CheckLinkResult.NO_SUCH_LINK.value
        print(f'Record for is {id} not found!')
        return JsonResponse({"result": result, "statusChanged": False, "delete": True})

    link = Obj.url
    pdf = Obj.pdfSource

    # does pdf still exist?
    if (os.path.isfile(pdf) == False):
        Obj.delete()
        result =  CheckLinkResult.NO_SUCH_PDF.value
        print(f'pdf:{pdf} not found!')
        return JsonResponse({"result": result, "statusChanged": False, "delete": True})

    # process the pdf to check if link still exists
    found, urlAnchor, moreInPdf = get_first_link_instance(link, pdf)
    if (found == False):
        Obj.delete()
        result =  CheckLinkResult.NO_SUCH_LINK.value
        print(f'url:{link} not found in pdf:{pdf}!')
        return JsonResponse({"result": result, "statusChanged": False, "delete": True})

    statusCode, reason, finalUrl = make_request(link)
 
    if (statusCode == 200):
        if (Obj.statusCode != statusCode):
            Obj.statusCode = statusCode
            Obj.broken = False
            Obj.dismiss = False
            Obj.ignore = False
            Obj.urlText = urlAnchor
            Obj.moreInPdf = moreInPdf
            Obj.save()
            return JsonResponse({"result": CheckLinkResult.LINK_OK.value, "statusChanged": True, "delete": False})
        return JsonResponse({"result": CheckLinkResult.LINK_OK.value, "statusChanged": False, "delete": False})
    else:
        if (Obj.statusCode != statusCode):
            Obj.dismiss = False
            Obj.broken = True
            Obj.statusCode = statusCode
            Obj.urlText = urlAnchor
            Obj.moreInPdf = moreInPdf
            Obj.save()
            result =  CheckLinkResult.LINK_BROKEN_STATUS_CHANGED.value
            return JsonResponse({"result": result, "statusChanged": True, "delete": False})
        result =  CheckLinkResult.LINK_BROKEN_SAME_STATUS.value
    # send response back with result 
    # result = one of (pdf does not exist, link does not exist, link ok, link broken same status, link broken status changed
    return JsonResponse({"result": result, "statusChanged": False, "delete": False})

""" def login_view(request):
    if request.method == "POST":

        # Attempt to sign user in
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)
        # Check if authentication successful
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse("index"))
        else:
            return render(request, "webcalendar/login.html", {
                "message": "Invalid username and/or password."
            })
    else:
        return render(request, "webcalendar/login.html") """

""" def register(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]

        # Ensure password matches confirmation
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        if password != confirmation:
            return HttpResponseRedirect("/register")
        # Attempt to create new user
        try:
            user = User.objects.create_user(username, email, password)
            user.save()

        except IntegrityError:
            return HttpResponseRedirect("/register")
            
        login(request, user)
        return HttpResponseRedirect("/")
    else:
        return HttpResponseRedirect("/register")       """

    
def register(request):
    if not request.user.is_superuser:
        return HttpResponseRedirect('/')
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Account created for {user.username}!')
            return redirect('profile')  # Redirect to profile after registration
    else:
        form = UserRegisterForm()
    return render(request, 'LinkChecker/register.html', {'form': form})      



def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return HttpResponseRedirect('/')  # Redirect to home or another page after successful login
        else:
            # Return an 'invalid login' error message.
            print('user authentication failed!')
            return render(request, 'LinkChecker/login.html', {'error': 'Invalid username or password.'})
    else:
        print('Login not a POST!')
        return render(request, 'LinkChecker/login.html')
@login_required
def manage(request):
    if not request.user.is_superuser:
        return HttpResponseRedirect('/')
    if request.method == 'GET':
        return render(request, 'LinkChecker/manage.html', {'users': User.objects.all()})
    if request.method == 'POST':
        userID = request.POST['id']
        User.objects.get(id=userID).delete()
        return HttpResponseRedirect('/manage')
@csrf_exempt
@login_required
def logout_view(request):
    logout(request)
    return HttpResponse(status=200)
    #return render(request, 'LinkChecker/logout.html')
    #return HttpResponseRedirect('logout') 
    #return HttpResponseRedirect(reverse('LinkChecker/index.html'))

@login_required
def profile(request):
    return render(request, 'LinkChecker/profile.html')
